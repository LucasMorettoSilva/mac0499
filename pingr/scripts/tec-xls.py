from openpyxl import Workbook
from openpyxl.styles import PatternFill


def read_file(filename):
    lines = []
    with open(filename) as file:
        for line in file:
            lines.append(line.split(","))
    return lines[1:]


def write(lines):
    wb = Workbook()
    sheet = wb.active

    red_fill = PatternFill("solid", start_color="FF0000")
    green_fill = PatternFill("solid", start_color="7FFF00")
    cyan_fill = PatternFill("solid", start_color="00B7EB")

    sheet.cell(row=1, column=1).value = "Language"
    sheet.cell(row=1, column=2).value = "AMQP Support (RabbitMQ)"
    sheet.cell(row=1, column=3).value = "AMQP Lib (RabbitMQ)"
    sheet.cell(row=1, column=4).value = "HTTP Support"
    sheet.cell(row=1, column=5).value = "HTTP Lib"

    for index, line in enumerate(lines, start=2):
        has_rabbit = bool(int(line[1]))
        has_http = bool(int(line[3]))

        sheet.cell(row=index, column=1).value = line[0]

        if index <= 13:
            sheet.cell(row=index, column=1).fill = cyan_fill

        sheet.cell(row=index, column=2).value = has_rabbit

        if has_rabbit:
            sheet.cell(row=index, column=2).fill = green_fill
            sheet.cell(row=index, column=3).hyperlink = line[2]
            sheet.cell(row=index, column=3).style = "Hyperlink"
            sheet.cell(row=index, column=3).value = "Link"
        else:
            sheet.cell(row=index, column=2).fill = red_fill

        sheet.cell(row=index, column=4).value = has_http

        if has_http:
            sheet.cell(row=index, column=4).fill = green_fill
            sheet.cell(row=index, column=5).hyperlink = line[4]
            sheet.cell(row=index, column=5).style = "Hyperlink"
            sheet.cell(row=index, column=5).value = "Link"
        else:
            sheet.cell(row=index, column=4).fill = red_fill

    wb.save("lang-support.xlsx")


def main():
    write(read_file("../../analysis/lang.txt"))


if __name__ == "__main__":
    main()
